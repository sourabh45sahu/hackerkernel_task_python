from django.shortcuts import render, redirect
from django.views.generic import ListView, CreateView
from django.urls import reverse_lazy
from .models import Author, Book, BorrowRecord
from .forms import AuthorForm, BookForm, BorrowRecordForm
import openpyxl
from django.http import HttpResponse

class AuthorListView(ListView):
    model = Author
    template_name = 'library_app/author_list.html'
    paginate_by = 10

class BookListView(ListView):
    model = Book
    template_name = 'library_app/book_list.html'
    paginate_by = 10

class BorrowRecordListView(ListView):
    model = BorrowRecord
    template_name = 'library_app/borrowrecord_list.html'
    paginate_by = 10

class AuthorCreateView(CreateView):
    model = Author
    form_class = AuthorForm
    template_name = 'library_app/author_form.html'
    success_url = reverse_lazy('author_list')

class BookCreateView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'library_app/book_form.html'
    success_url = reverse_lazy('book_list')

class BorrowRecordCreateView(CreateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = 'library_app/borrowrecord_form.html'
    success_url = reverse_lazy('borrowrecord_list')

def export_to_excel(request):
    authors = Author.objects.all()
    books = Book.objects.all()
    borrow_records = BorrowRecord.objects.all()

    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = 'Authors'

    sheet1.append(['ID', 'Name', 'Email', 'Bio'])
    for author in authors:
        sheet1.append([author.id, author.name, author.email, author.bio])

    sheet2 = workbook.create_sheet(title='Books')
    sheet2.append(['ID', 'Title', 'Genre', 'Published Date', 'Author'])
    for book in books:
        sheet2.append([book.id, book.title, book.genre, book.published_date, book.author.name])

    sheet3 = workbook.create_sheet(title='Borrow Records')
    sheet3.append(['ID', 'User Name', 'Book', 'Borrow Date', 'Return Date'])
    for record in borrow_records:
        sheet3.append([record.id, record.user_name, record.book.title, record.borrow_date, record.return_date])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'
    workbook.save(response)
    return response
